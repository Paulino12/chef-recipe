"""
Generate a supplier-first ingredient search catalog from Recipes Ingredients.xlsx.

Output:
  - web/data/ingredient_search_catalog.json
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
WORKBOOK_PATH = DATA_DIR / "Recipes Ingredients.xlsx"
OUTPUT_PATH = DATA_DIR / "ingredient_search_catalog.json"
WORKSHEET_NAME = "All Ingredients + Supplier"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_number(value: Any, digits: int = 8) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number, digits)


def normalize_bool(value: Any) -> bool:
    return normalize_text(value).lower() in {"yes", "true", "1"}


def detect_pricing_unit(pack_size: str, description: str, total_items: float | None) -> str:
    text = f"{pack_size} {description}".upper()

    if any(token in text for token in ("KG", " GR ", " G ", "GRAM", "G PK", "G EA")):
        return "G"
    if any(token in text for token in ("LTR", " L ", "ML", "CL")):
        return "ML"
    if total_items is not None and total_items.is_integer() and total_items <= 500:
        return "EA"
    return "EA"


def to_estimated_unit_price(price_per_item: float | None, pricing_unit: str) -> float | None:
    if price_per_item is None:
        return None
    if pricing_unit in {"G", "ML"}:
        return round(price_per_item / 1000, 8)
    return round(price_per_item, 8)


def build_searchable_text(parts: list[str]) -> str:
    combined = " ".join(part for part in parts if part)
    return re.sub(r"[^a-z0-9]+", " ", combined.lower()).strip()


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook[WORKSHEET_NAME]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
      raise RuntimeError("Workbook sheet is empty.")

    header = [normalize_text(value) for value in rows[0]]
    index = {name: position for position, name in enumerate(header)}

    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        supplier = normalize_text(row[index["Supplier"]])
        description = normalize_text(row[index["Description"]])
        if not supplier or not description:
            continue

        source_sheet = normalize_text(row[index["Source Sheet"]])
        major_description = normalize_text(row[index["Major Description"]])
        minor_description = normalize_text(row[index["Minor Description"]])
        manufacturer = normalize_text(row[index["Manufacturer"]])
        code = normalize_text(row[index["Code"]])
        pack_size = normalize_text(row[index["Pack Size"]])
        price = normalize_number(row[index["Price"]], 4)
        total_items = normalize_number(row[index["Total Items"]], 4)
        price_per_item = normalize_number(row[index["Price per Item"]], 8)
        weighted_item = normalize_bool(row[index["Weighted Item"]])
        pricing_unit = detect_pricing_unit(pack_size, description, total_items)
        estimated_unit_price = to_estimated_unit_price(price_per_item, pricing_unit)

        items.append(
            {
                "supplier": supplier,
                "source_sheet": source_sheet,
                "major_description": major_description,
                "minor_description": minor_description,
                "code": code,
                "description": description,
                "pack_size": pack_size,
                "price": price,
                "total_items": total_items,
                "price_per_item": price_per_item,
                "weighted_item": weighted_item,
                "manufacturer": manufacturer,
                "pricing_unit": pricing_unit,
                "estimated_unit_price": estimated_unit_price,
                "searchable_text": build_searchable_text(
                    [
                        description,
                        major_description,
                        minor_description,
                        manufacturer,
                        supplier,
                        code,
                    ]
                ),
            }
        )

    payload = {
        "generatedFromWorkbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sheet": WORKSHEET_NAME,
        "items": items,
    }
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(items)} ingredient search rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
